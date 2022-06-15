from selenium import webdriver
from selenium.webdriver import ChromeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome import service as fs
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver import Chrome

import time
import openpyxl


options = webdriver.ChromeOptions()
#ヘッドレスモードの場合指定
#options.add_argument('headless')
#シークレットモードの場合の指定
options.add_argument('incognito')
options.add_experimental_option("excludeSwitches", ['enable-automation']);
chrome_service = fs.Service(executable_path='C:\driver\chromedriver.exe')
serv = Service(ChromeDriverManager().install())#driverの自動更新
driver = webdriver.Chrome(service=serv, options=options)
wait = WebDriverWait(driver, 10)#タイムアウト時間の設定

#excelファイルを作成
# ブックを作成
book = openpyxl.Workbook()
# 保存する
book.save("shougakukin_list.xlsx")

#excelファイルを操作
# ブックを取得
book = openpyxl.load_workbook("shougakukin_list.xlsx")
# シートを取得
sheet = book['Sheet']
# セルへ書き込む
sheet.cell(row=1,column=1).value = "団体名"
sheet.cell(row=1,column=2).value = "給付開始日"
sheet.cell(row=1,column=3).value = "給付終了日"

# 保存する
book.save("shougakukin_list.xlsx")


# URLにアクセス
driver.get("https://xn--kus49bd41h.net/archives/67540341.html")

class_name = 'ex'                   # class属性名
class_elems = driver.find_elements_by_class_name(class_name) # classでの指定
 
# 取得した要素を1つずつ表示
i=2
for elem in class_elems:
    try:
        sheet.cell(row=i,column=1).value = elem.text
        print(elem.text)
        aTag    = elem.find_element_by_tag_name("a")
        url     = aTag.get_attribute("href")
        print(url)
        sheet.cell(row=i,column=1).hyperlink = url
        i+=1
    except:
        None

book.save("shougakukin_list.xlsx")
exit()
    

